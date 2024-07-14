import streamlit as st
import openpyxl
import re
import spacy
from io import BytesIO

# Load spaCy English model
nlp = spacy.load('en_core_web_sm')

def extract_vba_code_from_excel(file):
    wb = openpyxl.load_workbook(file, read_only=True)
    vba_code = ""

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and ("Sub " in cell or "Function " in cell):
                    vba_code += cell + "\n"

    return vba_code

def parse_vba_code(vba_code):
    variable_pattern = r'Dim ([^\n]*)'
    function_pattern = r'(?:Sub|Function) ([^\(]*)'
    comment_pattern = r"'.*"

    variables = re.findall(variable_pattern, vba_code)
    functions = re.findall(function_pattern, vba_code)
    comments = re.findall(comment_pattern, vba_code)

    return variables, functions, comments

def enhance_function_descriptions(functions):
    enhanced_functions = []
    for function in functions:
        doc = nlp(function)
        function_description = {
            'name': function,
            'entities': [ent.text for ent in doc.ents],  # Extract named entities
            'dependencies': [token.text for token in doc if token.dep_ == 'ROOT'],  # Extract root dependencies
        }
        enhanced_functions.append(function_description)

    return enhanced_functions

def main():
    st.title("VBA Macro Documentation and Analysis")
    st.write("""
    Upload an Excel file containing VBA macros to extract and analyze the code.
    The tool will provide details about the variables, functions, and comments found in the macros.
    """)

    uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

    if uploaded_file:
        st.write("Uploaded file details:")
        st.write(uploaded_file)

        # Process the uploaded file
        vba_code = extract_vba_code_from_excel(uploaded_file)
        variables, functions, comments = parse_vba_code(vba_code)

        st.write("Parsed Information:")
        st.write(f"Variables: {variables}")
        st.write(f"Functions: {functions}")
        st.write(f"Comments: {comments}")

        # Enhance function descriptions
        enhanced_functions = enhance_function_descriptions(functions)
        st.write("Function Descriptions (Enhanced with spaCy):")
        for desc in enhanced_functions:
            st.write(desc)

        # Create a downloadable document/text file
        download_text = f"""
        Parsed Information:
        Variables: {variables}
        Functions: {functions}
        Comments: {comments}

        Function Descriptions (Enhanced with spaCy):
        """
        for desc in enhanced_functions:
            download_text += f"\nName: {desc['name']}\nEntities: {desc['entities']}\nDependencies: {desc['dependencies']}\n"

        # Convert text to BytesIO object for download
        download_bytes = download_text.encode()
        st.download_button(
            label="Download Parsed Information",
            data=BytesIO(download_bytes),
            file_name='parsed_information.txt',
            mime='text/plain'
        )

if __name__ == "__main__":
    main()
